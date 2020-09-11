"""
General utility functions.
"""

import sys
import json
from pathlib import Path
from typing import Callable
import numpy as np
import talib
import openpyxl
from PyQt5 import QtGui

from .object import BarData, TickData
from .constant import Exchange, Interval


def pop_message_box(message):
    msgBox = QtGui.QMessageBox()
    msgBox.setWindowTitle(u'提示')
    msgBox.setText(message)
    # 隐藏ok按钮
    msgBox.addButton(QtGui.QMessageBox.Ok)
    msgBox.button(QtGui.QMessageBox.Ok).hide()
    # 模态对话框
    msgBox.exec_()


def extract_vt_symbol(vt_symbol: str):
    """
    :return: (symbol, exchange)
    """
    symbol, exchange_str = vt_symbol.split(".")
    return symbol, Exchange(exchange_str)


def generate_vt_symbol(symbol: str, exchange: Exchange):
    """
    return vt_symbol
    """
    return f"{symbol}.{exchange.value}"

# @Time    : 2019-10-10
# @Author  : Wang Yongchang
# 为pyinstaller打包exe,改变静态文件路径查找方式
# 此软件生成的临时文件夹.vntrader就在运行目录下
# 保证此软件是独立绿色软件，无需安装，
# 只需把pyinstaller打包后的文件夹放在空间大的磁盘中
# 运行依赖文件和临时生成文件都在运行目录下
# 拷贝和移动之后不影响之前保存的结果（回测等）
# 但是有丢失账户和脚本的风险，安全和方便二选一


def _get_trader_dir(temp_name: str):
    """
    Get path where trader is running in.
    """
    # cwd = Path.cwd()
    running_path = Path(sys.argv[0]).parent
    temp_path = running_path.joinpath(temp_name)

    # If .vntrader folder exists in current working directory,
    # then use it as trader running path.
    if temp_path.exists():
        return running_path, temp_path

    # Otherwise use home path of system.
    # home_path = Path.home()
    # temp_path = home_path.joinpath(temp_name)

    # Create .vntrader folder under home path if not exist.
    if not temp_path.exists():
        temp_path.mkdir()

    return running_path, temp_path


TRADER_DIR, TEMP_DIR = _get_trader_dir(".vntrader")


def get_file_path(filename: str):
    """
    Get path for temp file with filename.
    """
    return TEMP_DIR.joinpath(filename)


def get_folder_path(folder_name: str):
    """
    Get path for temp folder with folder name.
    """
    folder_path = TEMP_DIR.joinpath(folder_name)
    if not folder_path.exists():
        folder_path.mkdir()
    return folder_path

# @Time    : 2019-10-09
# @Author  : Wang Yongchang
# 为pyinstaller打包exe,改变静态文件路径查找方式
# 直接根据exe运行路径查找


def get_icon_path(ico_name: str):
    """
    Get path for icon file with ico name.
    """
    running_path = Path(sys.argv[0]).parent
    icon_path = running_path.joinpath("ico", ico_name)
    return str(icon_path)


def load_json(filename: str):
    """
    Load data from json file in temp path.
    """
    filepath = get_file_path(filename)

    if filepath.exists():
        with open(filepath, mode="r", encoding="UTF-8") as f:
            data = json.load(f)
        return data
    else:
        save_json(filename, {})
        return {}


def save_json(filename: str, data: dict):
    """
    Save data into json file in temp path.
    """
    filepath = get_file_path(filename)
    with open(filepath, mode="w+", encoding="UTF-8") as f:
        json.dump(
            data,
            f,
            indent=4,
            ensure_ascii=False
        )


def round_to(value: float, target: float):
    """
    Round price to price tick value.
    """
    rounded = int(round(value / target)) * target
    return rounded


def round_up(value: float):
    # 替换内置round函数,实现保留2位小数的精确四舍五入
    try:
        v = round(value * 100) / 100.0
    except OverflowError as e:
        v = '-'
    return v


class BarGenerator:
    """
    For: 
    1. generating 1 minute bar data from tick data
    2. generateing x minute bar/x hour bar data from 1 minute data

    Notice:
    1. for x minute bar, x must be able to divide 60: 2, 3, 5, 6, 10, 15, 20, 30
    2. for x hour bar, x can be any number
    """

    def __init__(
        self,
        on_bar: Callable,
        window: int = 0,
        on_window_bar: Callable = None,
        interval: Interval = Interval.MINUTE
    ):
        """Constructor"""
        self.bar = None
        self.on_bar = on_bar

        self.interval = interval
        self.interval_count = 0

        self.window = window
        self.window_bar = None
        self.on_window_bar = on_window_bar

        self.last_tick = None
        self.last_bar = None

    def update_tick(self, tick: TickData):
        """
        Update new tick data into generator.
        """
        new_minute = False

        # Filter tick data with 0 last price
        if not tick.last_price:
            return

        if not self.bar:
            new_minute = True
        elif self.bar.datetime.minute != tick.datetime.minute:
            self.bar.datetime = self.bar.datetime.replace(
                second=0, microsecond=0
            )
            self.on_bar(self.bar)

            new_minute = True

        if new_minute:
            self.bar = BarData(
                symbol=tick.symbol,
                exchange=tick.exchange,
                interval=Interval.MINUTE,
                datetime=tick.datetime,
                gateway_name=tick.gateway_name,
                open_price=tick.last_price,
                high_price=tick.last_price,
                low_price=tick.last_price,
                close_price=tick.last_price,
                open_interest=tick.open_interest
            )
        else:
            self.bar.high_price = max(self.bar.high_price, tick.last_price)
            self.bar.low_price = min(self.bar.low_price, tick.last_price)
            self.bar.close_price = tick.last_price
            self.bar.open_interest = tick.open_interest
            self.bar.datetime = tick.datetime

        if self.last_tick:
            volume_change = tick.volume - self.last_tick.volume
            self.bar.volume += max(volume_change, 0)

        self.last_tick = tick

    def update_bar(self, bar: BarData):
        """
        Update 1 minute bar into generator
        """
        # If not inited, creaate window bar object
        if not self.window_bar:
            # Generate timestamp for bar data
            if self.interval == Interval.MINUTE:
                dt = bar.datetime.replace(second=0, microsecond=0)
            else:
                dt = bar.datetime.replace(minute=0, second=0, microsecond=0)

            self.window_bar = BarData(
                symbol=bar.symbol,
                exchange=bar.exchange,
                datetime=dt,
                gateway_name=bar.gateway_name,
                open_price=bar.open_price,
                high_price=bar.high_price,
                low_price=bar.low_price
            )
        # Otherwise, update high/low price into window bar
        else:
            self.window_bar.high_price = max(
                self.window_bar.high_price, bar.high_price)
            self.window_bar.low_price = min(
                self.window_bar.low_price, bar.low_price)

        # Update close price/volume into window bar
        self.window_bar.close_price = bar.close_price
        self.window_bar.volume += int(bar.volume)
        self.window_bar.open_interest = bar.open_interest

        # Check if window bar completed
        finished = False

        if self.interval == Interval.MINUTE:
            # x-minute bar
            if not (bar.datetime.minute + 1) % self.window:
                finished = True
        elif self.interval == Interval.HOUR:
            if self.last_bar and bar.datetime.hour != self.last_bar.datetime.hour:
                # 1-hour bar
                if self.window == 1:
                    finished = True
                # x-hour bar
                else:
                    self.interval_count += 1

                    if not self.interval_count % self.window:
                        finished = True
                        self.interval_count = 0

        if finished:
            self.on_window_bar(self.window_bar)
            self.window_bar = None

        # Cache last bar object
        self.last_bar = bar

    def generate(self):
        """
        Generate the bar data and call callback immediately.
        """
        self.bar.datetime = self.bar.datetime.replace(
            second=0, microsecond=0
        )
        self.on_bar(self.bar)
        self.bar = None


class ArrayManager(object):
    """
    For:
    1. time series container of bar data
    2. calculating technical indicator value
    """

    def __init__(self, size=300):
        """Constructor"""
        self.count = 0
        self.size = size
        self.inited = False

        self.open_array = np.zeros(size)
        self.high_array = np.zeros(size)
        self.low_array = np.zeros(size)
        self.close_array = np.zeros(size)
        self.volume_array = np.zeros(size)

    def update_bar(self, bar):
        """
        Update new bar data into array manager.
        """
        self.count += 1
        if not self.inited and self.count >= self.size:
            self.inited = True

        self.open_array[:-1] = self.open_array[1:]
        self.high_array[:-1] = self.high_array[1:]
        self.low_array[:-1] = self.low_array[1:]
        self.close_array[:-1] = self.close_array[1:]
        self.volume_array[:-1] = self.volume_array[1:]

        self.open_array[-1] = bar.open_price

        self.high_array[-1] = bar.high_price
        self.low_array[-1] = bar.low_price
        self.close_array[-1] = bar.close_price
        self.volume_array[-1] = bar.volume

    @property
    def open(self):
        """
        Get open price time series.
        """
        return self.open_array

    @property
    def high(self):
        """
        Get high price time series.
        """
        return self.high_array

    @property
    def low(self):
        """
        Get low price time series.
        """
        return self.low_array

    @property
    def close(self):
        """
        Get close price time series.
        """
        return self.close_array

    @property
    def volume(self):
        """
        Get trading volume time series.
        """
        return self.volume_array

    # def sma(self, n, array=False):
    #     """
    #     Simple moving average.
    #     """
    #     # print(self.close)
    #     result = talib.SMA(self.close, n)
    #     if array:
    #         return result
    #     return result[-1]

    def sma(self, flag, n, array=False):
        """
        Simple moving average.
        """
        # if len(flag) > 1:
        #     ohlc = {'OPEN':self.open,'HIGH':self.high,'LOW':self.low,'CLOSE':self.close}
        # else:
        ohlc = {'o': self.open, 'h': self.high, 'l': self.low, 'c': self.close}
        result = talib.SMA(ohlc[flag], n)
        if array:
            return result
        return result[-1]

    def std(self, n, array=False):
        """
        Standard deviation
        """
        result = talib.STDDEV(self.close, n)
        if array:
            return result
        return result[-1]

    def cci(self, n, array=False):
        """
        Commodity Channel Index (CCI).
        """
        result = talib.CCI(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def atr(self, n, array=False):
        """
        Average True Range (ATR).
        """
        result = talib.ATR(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def rsi(self, n, array=False):
        """
        Relative Strenght Index (RSI).
        """
        result = talib.RSI(self.close, n)
        if array:
            return result
        return result[-1]

    def macd(self, fast_period, slow_period, signal_period, array=False):
        """
        MACD.
        """
        macd, signal, hist = talib.MACD(
            self.close, fast_period, slow_period, signal_period
        )
        if array:
            return macd, signal, hist
        return macd[-1], signal[-1], hist[-1]

    def adx(self, n, array=False):
        """
        ADX.
        """
        result = talib.ADX(self.high, self.low, self.close, n)
        if array:
            return result
        return result[-1]

    def boll(self, n, dev, array=False):
        """
        Bollinger Channel.
        """
        mid = self.sma("c", n, array)
        std = self.std(n, array)

        up = mid + std * dev
        down = mid - std * dev

        return up, down

    def keltner(self, n, dev, array=False):
        """
        Keltner Channel.
        """
        mid = self.sma("c", n, array)
        atr = self.atr(n, array)

        up = mid + atr * dev
        down = mid - atr * dev

        return up, down

    def donchian(self, n, array=False):
        """
        Donchian Channel.
        """
        up = talib.MAX(self.high, n)
        down = talib.MIN(self.low, n)

        if array:
            return up, down
        return up[-1], down[-1]


class FunctionLoader(object):
    def __init__(self):
        self.names = []
        self.contents = {}

    def load_excel(self, path):
        workbook = openpyxl.load_workbook(path)
        # 获取表单对象
        try:
            sheet = workbook.get_sheet_by_name('Sheet1')
        except Exception:
            pass

        # 获取行数
        nrow = sheet.max_row
        # 获取列数
        # 遍历读取
        for row in range(2, nrow + 1):
            item = {}
            # 设置值
            item["name"] = sheet.cell(row=row, column=1).value
            item["required"] = bool(sheet.cell(row=row, column=2).value)
            item["status"] = sheet.cell(row=row, column=3).value
            item["description"] = sheet.cell(row=row, column=4).value
            item["note"] = sheet.cell(row=row, column=5).value
            item["example"] = sheet.cell(row=row, column=6).value
            item["condition"] = sheet.cell(row=row, column=7).value

            # 添加到函数字典中
            if item["required"]:
                self.names.append(item["name"])
                self.contents[item["name"]] = item

class CTPLoader(object):
    def __init__(self):
        self.names = []
        self.contents = {}

    def load_excel(self, path):
        workbook = openpyxl.load_workbook(path)
        # 获取表单对象
        try:
            sheet = workbook.get_sheet_by_name('Sheet4')
        except Exception:
            pass

        # 获取行数
        nrow = sheet.max_row
        # 获取列数
        # 遍历读取
        for row in range(2, nrow + 1):
            item = {}
            # 设置值
            item["InstrumentID"] = sheet.cell(row=row, column=1).value
            item["ExchangeID"] = bool(sheet.cell(row=row, column=2).value)
            item["InstrumentName"] = sheet.cell(row=row, column=3).value
            item["ExchangeInstID"] = sheet.cell(row=row, column=4).value
            item["ProductID"] = sheet.cell(row=row, column=5).value
            item["VolumeMultiple"] = sheet.cell(row=row, column=6).value
            item["PriceTick"] = sheet.cell(row=row, column=7).value

            # 添加到函数字典中
            #if item["required"]:
            self.names.append(item["InstrumentID"])
            self.contents[item["InstrumentID"]] = item
def virtual(func: "callable"):
    """
    mark a function as "virtual", which means that this function can be override.
    any base class should use this or @abstractmethod to decorate all functions
    that can be (re)implemented by subclasses.
    """
    return func

