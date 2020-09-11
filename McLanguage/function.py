import openpyxl


class Function(object):
    def __init__(self):
        # 函数名
        self.name = ''
        # 函数参数是否字符串化
        self.param_stringizing = False
        # 函数类型
        self.type = ''
        # 函数功能描述
        self.describe = ''
        # 函数使用注意要项
        self.note = ''
        # 函数使用例子
        self.example = ''

# 函数字典
functions = {}
# 读取函数描述文件
# 获取工作簿
workbook = openpyxl.load_workbook('function.xlsx')
# 获取表单对象
try:
    sheet = workbook.get_sheet_by_name('Sheet1')
except Exception:
    pass
# 获取行数
nrow = sheet.max_row
# 获取列数
ncol = sheet.max_column
# 遍历读取
for row in range(2, nrow+1):
    func = Function()
    # 设置值
    func.name = sheet.cell(row=row, column=1).value
    func.param_stringizing = bool(sheet.cell(row=row, column=2).value)
    func.type = sheet.cell(row=row, column=3).value
    func.describe = sheet.cell(row=row, column=4).value
    func.note = sheet.cell(row=row, column=5).value
    func.example = sheet.cell(row=row, column=6).value
    # 添加到函数字典中
    functions[func.name] = func

if __name__ == '__main__':
    print("函数名 \t 字符化 \t 类型 \t 功能描述 \t 注释 \t 样例")
    for func in functions.values():
        print("{!r} \t {!r} \t {!r} \t {!r} \t {!r} \t {!r}".format(
            func.name, func.param_stringizing, func.type, func.describe, func.note, func.example))